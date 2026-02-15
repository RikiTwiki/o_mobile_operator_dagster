from __future__ import annotations

import os
from io import BytesIO
from typing import Any, Dict, List, Optional

import requests
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from dagster import op, get_dagster_logger, Failure, Out, In

from resources import ReportUtils
from utils.array_operations import left_join_array
from utils.getters.mp_occupancy_data_getter import MPOccupancyDataGetter
from utils.path import BASE_SQL_PATH

def _parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    s = str(value).strip()
    s10 = s[:10]  # режем "YYYY-MM-DD ..." / "DD.MM.YYYY ..."
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s10, fmt).date()
        except ValueError:
            pass
    return None

@op(required_resource_keys={'source_dwh', 'source_naumen_3', 'source_naumen_1', 'report_utils'})
def get_chats_op(context):
    dates = context.resources.report_utils.get_utils()
    start_date = dates.StartDate
    end_date = dates.EndDate
    date_list = dates.Dates
    report_date = dates.ReportDate
    month_start = dates.MonthStart
    month_end = dates.MonthEnd
    formatted_dates = dates.FormattedDates

    conn = context.resources.source_dwh
    sql_path = Path(BASE_SQL_PATH) / "stat" / "mp_handling_time.sql"
    sql_template = sql_path.read_text(encoding="utf-8")
    params = {
        "start_date": report_date,
        "end_date": end_date,
    }
    df = pd.read_sql(sql_template, conn, params=params)
    chats = df.to_dict(orient='records')

    mp_state_getter = MPOccupancyDataGetter(start_date=report_date, end_date=end_date, conn=conn, trunc='day')
    mp_state_getter.month_start = datetime.strptime(report_date, "%Y-%m-%d").replace(day=1).strftime("%Y-%m-01")
    mp_state_getter.set_days_trunked_data()

    data = mp_state_getter.get_active_user_logins()

    result = left_join_array(chats, data, 'login')
    return result


@op(required_resource_keys={'source_dwh', 'source_naumen_3', 'source_naumen_1', 'report_utils', 'source_cp'})
def get_user_statuses_op(context):
    dates = context.resources.report_utils.get_utils()
    start_date = dates.StartDate
    end_date = dates.EndDate
    date_list = dates.Dates
    report_date = dates.ReportDate
    month_start = dates.MonthStart
    month_end = dates.MonthEnd
    formatted_dates = dates.FormattedDates

    occupancy_data_getter = MPOccupancyDataGetter(conn=context.resources.source_dwh, trunc='day')
    occupancy_data_getter.pst_positions = [10, 41, 46]
    occupancy_data_getter.start_date = report_date
    occupancy_data_getter.end_date = end_date
    occupancy_data_getter.set_days_trunked_data()
    occupancy_data_getter.month_start = month_start

    cross_day_users = occupancy_data_getter.get_active_user_logins([9, 10, 12, 13, 19])
    cross_day_logins = {row.get("login") for row in cross_day_users if row.get("login")}

    all_users = occupancy_data_getter.get_active_user_logins()
    all_logins = [row.get("login") for row in all_users if row.get("login")]
    normal_shift_logins = [lg for lg in all_logins if lg not in cross_day_logins]

    cross_day_shift_sheets = occupancy_data_getter.get_active_user_sheets()

    occupancy_data_getter = MPOccupancyDataGetter(conn=context.resources.source_cp, trunc='day')
    occupancy_data_getter.pst_positions = [10, 41, 46]
    occupancy_data_getter.start_date = report_date
    occupancy_data_getter.end_date = end_date
    occupancy_data_getter.set_days_trunked_data()
    occupancy_data_getter.month_start = month_start

    normal_shift_statuses = occupancy_data_getter.get_user_statuses_by_positions(normal_shift_logins)
    cross_day_shift_statuses = occupancy_data_getter.get_sheet_statuses_by_positions(cross_day_shift_sheets)

    groups = defaultdict(list)
    for it in cross_day_shift_statuses:
        groups[(it.get("date"), it.get("login"), it.get("code"))].append(it)

    grouped_cross = []
    for (date, login, code), grp in groups.items():
        seconds = sum((x.get("seconds") or 0) for x in grp)
        minutes = round(sum(float(x.get("minutes") or 0) for x in grp), 2)
        hours   = round(sum(float(x.get("hours")   or 0) for x in grp), 3)
        grouped_cross.append({"date": date, "login": login, "code": code,
                              "seconds": seconds, "minutes": minutes, "hours": hours})

    statuses = list(normal_shift_statuses) + grouped_cross
    user_by_login = {u.get("login"): u for u in all_users}
    result = []
    for row in statuses:
        u = user_by_login.get(row.get("login"), {})
        result.append({**row, **{k: v for k, v in u.items() if k != "login"}})
    return result

def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _set_widths(ws, widths: Dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _build_excel_chats(chats: List[Dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Длительность обработки чатов"
    header = ['Дата', 'chat_id', 'Группа', 'Логин', 'Супервайзер',
              'Время первой реакции', 'Среднее время реакции',
              'Количество ответов', 'номер распределения']
    ws.append(header)
    ws.freeze_panes = "A2"
    _set_widths(ws, {"A":20,"B":7,"C":7,"D":20,"E":20,"F":22,"G":24,"H":20,"I":22})

    for r in chats:
        d = _parse_date(r.get("date"))
        ws.append([
            d,  # <-- пишем date, не строку
            r.get("chat_id",""), r.get("abbreviation",""),
            r.get("login",""), r.get("supervisor",""),
            r.get("user_reaction_time",""), r.get("user_average_replies_time",""),
            r.get("user_count_replies",""), r.get("agent_leg_number",""),
        ])
        if d:
            ws.cell(row=ws.max_row, column=1).number_format = "dd-mm-yyyy"

    return wb

def _build_excel_statuses(statuses: List[Dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "ГПО | Статусы пользователей"
    header = ['Дата','Группа','Логин','Супервайзер','Статус',
              'Длительность(сек)','Длительность(мин)','Длительность(часы)']
    ws.append(header)
    ws.freeze_panes = "A2"
    _set_widths(ws, {"A":20,"B":7,"C":20,"D":20,"E":12,"F":18,"G":18,"H":18})

    for r in statuses:
        d = _parse_date(r.get("date"))
        ws.append([
            d,  # <-- date
            r.get("abbreviation",""), r.get("login",""),
            r.get("supervisor",""), r.get("code",""),
            r.get("seconds",""), r.get("minutes",""), r.get("hours",""),
        ])
        if d:
            ws.cell(row=ws.max_row, column=1).number_format = "dd-mm-yyyy"

    return wb


@op
def text_handling_times_excel(chats, statuses):
    wb1 = _build_excel_chats(chats or [])
    wb2 = _build_excel_statuses(statuses or [])

    Path("./reports/mp_text").mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    files = {
        f"rt_over_15_min_{stamp}.xlsx": _wb_to_bytes(wb1),
        f"user_statues_{stamp}.xlsx": _wb_to_bytes(wb2),
    }

    saved: dict[str, str] = {}
    for name, data in files.items():
        p = Path("./reports/mp_text") / name
        p.write_bytes(data)
        saved[name] = str(p)
    return saved

@op(
    required_resource_keys={"report_utils"},
    ins={"saved_files": In(dict, description="dict вида {filename: path} из text_handling_times_excel")},
    out=Out(dict, description="Результаты загрузки по каждому файлу"),
)
def upload_text_handling_times_excels_op(context, saved_files: Dict[str, str]) -> Dict[str, Any]:
    cfg = context.op_config or {}

    delete_after = bool(cfg.get("delete_after_upload", True))
    delete_only_on_2xx = bool(cfg.get("delete_only_on_2xx", True))

    # day только из report_utils
    dates = context.resources.report_utils.get_utils()
    rd = dates.ReportDate
    if isinstance(rd, str):
        report_day = datetime.strptime(rd[:10], "%Y-%m-%d").date()
    elif isinstance(rd, datetime):
        report_day = rd.date()
    else:
        report_day = rd  # date

    day_str = report_day.strftime("%Y.%m.%d")

    base_url = cfg.get("base_url") or os.getenv("REPORTS_API_BASE", "https://rv.o.kg")
    report_id = int(cfg.get("report_id", 12))
    api_key = cfg.get("api_key") or os.getenv("REPORTS_API_KEY") or os.getenv("RV_REPORTS_API_KEY")
    if not api_key:
        raise Failure(
            description="Не задан API ключ. Укажите op_config.api_key или REPORTS_API_KEY / RV_REPORTS_API_KEY."
        )

    timeout_connect = int(cfg.get("timeout_connect", 10))
    timeout_read = int(cfg.get("timeout_read", 300))
    verify_ssl = bool(cfg.get("verify_ssl", True))

    title_base = cfg.get("title") or "ТЕКСТОВЫЕ | Статистика для мониторинга превышений ГПО"

    url = f"{base_url.rstrip('/')}/api/report-instances/upload/{report_id}"
    headers = {"x-api-key": api_key}

    context.log.info(f"Upload day={day_str}, url={url}, verify_ssl={verify_ssl}")

    results: Dict[str, Any] = {"day": day_str, "report_id": report_id, "uploaded": {}, "errors": {}}

    if not saved_files:
        context.log.warning("saved_files is empty — nothing to upload")
        return results

    for filename, file_path in saved_files.items():
        try:
            path = str(file_path)
            if not os.path.exists(path):
                raise FileNotFoundError(f"File not found: {path}")

            data = {
                "title": f"{title_base} | {filename}",
                "day": day_str,
            }

            context.log.info(f"POST {url} file={path}")

            with open(path, "rb") as f:
                files = {
                    "file": (
                        os.path.basename(path),
                        f,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                }
                resp = requests.post(
                    url,
                    headers=headers,
                    data=data,
                    files=files,
                    timeout=(timeout_connect, timeout_read),
                    verify=verify_ssl,
                )

            context.log.info(f"{filename}: status={resp.status_code}")

            try:
                payload = resp.json()
            except Exception:
                payload = {"status_code": resp.status_code, "text": resp.text}

            # если API вернул ошибку кодом — всё равно сохраним ответ
            results["uploaded"][filename] = payload

            success = 200 <= resp.status_code < 300

            if delete_after and os.path.exists(path):
                if (not delete_only_on_2xx) or success:
                    os.remove(path)
                    context.log.info(f"{filename}: deleted local file {path}")
                else:
                    context.log.info(
                        f"{filename}: not deleted (status={resp.status_code}, delete_only_on_2xx={delete_only_on_2xx})")

        except Exception as e:
            context.log.error(f"{filename}: upload error: {e}")
            results["errors"][filename] = {"status_code": -1, "text": str(e)}

    return results